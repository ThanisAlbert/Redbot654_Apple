from tkinter import messagebox
import time
import openpyxl
from selenium.webdriver.support.wait import WebDriverWait
from ChromeDriver.Chromedriver import Chromedriver
from Log.Log import Log
from Web.WebNavigate import WebNavigate
from Web.WebWait import WebWait

class ProcessExcel(WebWait,Log):

    def __init__(self,inputfile,label_status):
        self.log = self.getLogger()
        self.inputfile =inputfile
        self.chrome = Chromedriver(downloadpath="d:\\", headless=False)
        self.driver = self.chrome.chromedriver()
        self.label_status =label_status
        self.wait = WebDriverWait(self.driver, 40)
        WebWait.__init__(self, self.wait)

    def process_excel(self):

        apple_wk = openpyxl.load_workbook(self.inputfile,keep_vba=True)
        apple_sht= apple_wk['Sheet1']

        maxrow = 0
        for i in range(1, 10000):
            if str(apple_sht.cell(row=i, column=1).value).strip() != "None" and str(apple_sht.cell(row=i, column=1).value).strip() != "" :
                maxrow = maxrow+1

        for i in range(2, maxrow + 1):
            apple_sht.cell(row=i, column=41).value = ""
            apple_sht.cell(row=i, column=42).value = ""
            apple_sht.cell(row=i, column=43).value = ""
            apple_sht.cell(row=i, column=44).value = ""
            apple_sht.cell(row=i, column=45).value = ""
            apple_sht.cell(row=i, column=46).value = ""
            apple_sht.cell(row=i, column=47).value = ""

        apple_wk.save(self.inputfile)

        for i in range(2,maxrow+1):
            try:
                self.label_status.configure(text=str("Downloading ") + str(i - 1) + str(" of ") + str(maxrow - 1))
                self.label_status.update()
                id = apple_sht.cell(row=i, column=1).value
                vendorcode = apple_sht.cell(row=i, column=16).value
                description = apple_sht.cell(row=i, column=15).value
                location = apple_sht.cell(row=i, column=23).value
                if "Mumbai" in location:
                    location="BHIWANDI"
                apple_sht.cell(row=i, column=47).value = location
                quantity = apple_sht.cell(row=i, column=11).value
                ponumber = apple_sht.cell(row=i, column=35).value

                webnavigate = WebNavigate(applesht = apple_sht, row=i, driver=self.driver, ponum=ponumber, vendorcode=vendorcode, description=description, location=location,qty=quantity)
                webnavigate.web_navigate()

                self.wait_for_element_byxpath("//*[contains(text(),'Place Order')]")
                finalplaceorder = self.driver.find_element_by_xpath("//*[contains(text(),'Place Order')]")
                self.wait_for_element_byxpath("//div[contains(text(), 'By placing an order you acknowledge that you are subject to')]")

                self.wait_for_element_byxpath("//td[@headers='Unit Price']/span[@role='text']")
                pounitprice = self.driver.find_element_by_xpath("//td[@headers='Unit Price']/span[@role='text']").text

                self.wait_for_element_byxpath("//td[@headers='Total Price']/span[@role='text']")
                pototalprice = self.driver.find_element_by_xpath("//td[@headers='Total Price']/span[@role='text']").text

                self.wait_for_element_byxpath("//td[@headers='totalquantity']")
                poqty = self.driver.find_element_by_xpath("//td[@headers='totalquantity']").text
                self.log.info("unitprice, totalprice, qty taken")

                specificationtext = ""
                self.wait_for_element_byxpath("//button[@class='btn btn-link specifications']")
                specification = self.driver.find_element_by_xpath("//button[@class='btn btn-link specifications']")
                specification.click()
                self.log.info("specification clicked")
                time.sleep(2.5)
                self.wait_for_element_byxpath("//div[@id='specification0']/div/ul")
                specification_ul = self.driver.find_element_by_xpath("//div[@id='specification0']/div/ul")
                items = specification_ul.find_elements_by_tag_name("li")
                for item in items:
                    specificationtext = specificationtext + "//" + item.text
                self.log.info("specification taken")

                pounitprice = str(pounitprice).replace("INR","").replace(",","").strip()
                pounitprice = float(pounitprice)

                apple_sht.cell(row=i, column=25).value = pounitprice
                apple_sht.cell(row=i, column=41).value = pounitprice
                apple_sht.cell(row=i, column=42).value = poqty
                apple_sht.cell(row=i, column=43).value = pototalprice
                apple_sht.cell(row=i, column=45).value = specificationtext
                apple_wk.save(self.inputfile)

                self.log.info("OrderPlacedSuccessfully")
                self.log.info("======================================="+ str(id) +"=======================================================")

            except Exception as e:
                self.driver.close()
                self.log.info(e)
                apple_sht.cell(row=i, column=41).value = "error"
                apple_sht.cell(row=i, column=42).value = "error"
                apple_sht.cell(row=i, column=43).value = "error"
                apple_sht.cell(row=i, column=45).value = "error"
                apple_wk.save(self.inputfile)
                self.log.info("Error Order Placed")
                self.log.info("======================================="+ str(id) +"=============================================")
                self.chrome = Chromedriver(downloadpath="d:\\", headless=False)
                self.driver = self.chrome.chromedriver()
                self.wait = WebDriverWait(self.driver, 30)
                WebWait.__init__(self, self.wait)

        apple_wk.close()
        messagebox.showinfo("showinfo", "Completed")



